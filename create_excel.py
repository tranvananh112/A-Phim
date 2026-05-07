import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_excel():
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: BẢNG GIÁ & HỢP ĐỒNG ---
    ws1 = wb.active
    ws1.title = "Hợp Đồng Quảng Cáo"
    
    # Colors
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    sub_header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    white_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    
    # Title
    ws1.merge_cells('A1:G1')
    cell = ws1['A1']
    cell.value = "HỢP ĐỒNG QUẢNG CÁO A PHIM PREMIUM"
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30
    
    # General Info
    ws1['A3'] = "Người cung cấp dịch vụ:"
    ws1['B3'] = "Website Xem Phim (Bên A)"
    ws1['D3'] = "Website: https://aphim.io.vn/"
    
    ws1['A4'] = "Đại diện đối tác (Bên B):"
    ws1['B4'] = "ADMIN : KhangZenit"
    ws1['D4'] = "Website: https://gemclub.pro"
    
    ws1['A6'] = "Ngày lập:"
    ws1['B6'] = "16/05/2026"
    
    ws1['A7'] = "Trạng thái hợp đồng:"
    ws1['B7'] = "⏳ Chờ xác nhận và tiến hành lắp đặt"
    
    # Traffic Info
    ws1.merge_cells('A9:G9')
    ws1['A9'] = "🌟 THÔNG TIN TRAFFIC HỆ THỐNG A PHIM (Tháng 4/2026)"
    ws1['A9'].font = bold_font
    ws1['A9'].fill = sub_header_fill
    
    ws1['A10'] = "👥 Users / Tháng:"
    ws1['B10'] = "128,000 Users"
    ws1['A11'] = "⚡ Realtime Peak:"
    ws1['B11'] = "1,847 Online"
    ws1['A12'] = "⏱ Avg. Session:"
    ws1['B12'] = "9 phút 24 giây"
    
    for row in range(3, 8):
        ws1[f'A{row}'].font = bold_font
    for row in range(10, 13):
        ws1[f'A{row}'].font = bold_font
        ws1[f'B{row}'].font = bold_font
    
    # --- Section I ---
    ws1.merge_cells('A14:G14')
    ws1['A14'] = "I. CHI TIẾT GÓI PREMIUM – COMBO C (8 VỊ TRÍ QUẢNG CÁO)"
    ws1['A14'].font = bold_font
    ws1['A14'].fill = sub_header_fill
    
    headers_1 = ["STT", "Loại Quảng Cáo", "Đơn Giá Gốc/Ngày", "Đơn Giá/Ngày (Đã giảm 20%)", "Thu/Tuần (VNĐ)", "Thu/Tháng (VNĐ)", "Ghi Chú"]
    for col_num, header in enumerate(headers_1, 1):
        cell = ws1.cell(row=15, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style='thin'))
    
    data_premium = [
        [1, "Toàn màn hình", 187500, 150000, 1050000, 4500000, "Vị trí chính – nổi bật nhất"],
        [2, "Banner tĩnh (3 vị trí)", 225000, 180000, 1260000, 5400000, "Header + Sidebar + Footer"],
        [3, "Banner động", 150000, 120000, 840000, 3600000, "Hiển thị động"],
        [4, "Pop-under", 100000, 80000, 560000, 2400000, "Mở tab phía sau trình duyệt"],
        [5, "Native Ad", 312500, 250000, 1750000, 7500000, "Tích hợp trong danh sách phim"]
    ]
    
    for row_num, row_data in enumerate(data_premium, 16):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = cell_data
            if col_num in [3, 4, 5, 6]:
                cell.number_format = '#,##0'
                
    # Totals
    ws1['B21'] = "TỔNG CỘNG (8 VỊ TRÍ):"
    ws1['B21'].font = bold_font
    
    ws1['C21'] = 975000
    ws1['D21'] = 780000
    ws1['E21'] = 5460000
    ws1['F21'] = 23400000
    
    for col in ['C', 'D', 'E', 'F']:
        ws1[f'{col}21'].font = bold_font
        ws1[f'{col}21'].number_format = '#,##0'
        ws1[f'{col}21'].fill = highlight_fill

    ws1.merge_cells('A23:G23')
    ws1['A23'] = "★ Giá giảm 20% dành riêng cho đối tác ký hợp đồng trước 10/05/2026 ★"
    ws1['A23'].font = Font(italic=True, bold=True, color="FF0000")
    ws1['A23'].alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 30
    
    # --- Sheet 2: Lịch Thanh Toán ---
    ws2 = wb.create_sheet(title="Lịch Thanh Toán")
    
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "LỊCH THANH TOÁN DỰ KIẾN (GÓI PREMIUM)"
    ws2['A1'].font = Font(color="FFFFFF", bold=True, size=16)
    ws2['A1'].fill = header_fill
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30
    
    headers_2 = ["Kỳ Thanh Toán", "Hạng Mục", "Số Tiền (VNĐ)", "Trạng Thái", "Ghi Chú"]
    for col_num, header in enumerate(headers_2, 1):
        cell = ws2.cell(row=3, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style='thin'))
        
    data_payment = [
        ["Khi lắp xong – Tuần 1", "Phí Premium tuần 1 (Đã giảm 20%)", 5460000, "⏳ CHỜ THU", "780.000 × 7 ngày"],
        ["Tuần 2", "Phí Premium tuần 2 (Đã giảm 20%)", 5460000, "🔄 ĐỊNH KỲ/TUẦN", ""],
        ["Tuần 3", "Phí Premium tuần 3 (Đã giảm 20%)", 5460000, "🔄 ĐỊNH KỲ/TUẦN", ""],
        ["Tuần 4", "Phí Premium tuần 4 (Đã giảm 20%)", 5460000, "🔄 ĐỊNH KỲ/TUẦN", ""],
        ["Thanh Toán Theo Tháng", "Hoặc thu nguyên tháng (Đã giảm 20%)", 23400000, "📋 TÙY CHỌN", "Thanh toán 1 lần cho 30 ngày"]
    ]
    
    for row_num, row_data in enumerate(data_payment, 4):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.value = cell_data
            if col_num == 3:
                cell.number_format = '#,##0'
                
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 35
    
    # Save the file
    wb.save("Hop_Dong_Quang_Cao_A_Phim_x_GemClub.xlsx")

create_excel()
